"""THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHOR BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN 
ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION 
WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE."""

import streamlit as st

import numpy as np
import pandas as pd

import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

import re
import sys

st.set_page_config(page_title="DCB",page_icon="⏩")

m = st.markdown("""
<style>
div.stButton > button:first-child {
    font-size:16px;font-weight:bold;height:2em;width:7em;
}
</style>""", unsafe_allow_html=True)


st.title('DCB Data Analysis')


instron_file_ref=st.file_uploader("Choose DCB Raw Data CSV file  | Force according to selection & Displacement in mm")

run_button=st.button("Run")

if run_button:

    def arranjar_dcb (instron_csv_file):#Function - Input: Ficheiro RAW da Instron / Atenção ao nome das Colunas (!!); Output: DataFrame Organizado Instron

        df1_instron = pd.read_csv(instron_csv_file,sep=",",usecols= [1,2], names=["Displacement","Force"],header=8)

        return df1_instron


    tabela_final=arranjar_dcb (instron_file_ref)


    st.dataframe(tabela_final)

    
    fig = px.scatter(tabela_final, x='Eyy', y='Tensile Stress', marginal_y="box",
           marginal_x="box",template="ggplot2")
    fig.update_layout(
        yaxis = dict(
            tickmode = 'linear',
            tick0 = 0,
            dtick = 500,
            tickformat = '.2f'
        )
    )
    st.plotly_chart(fig, use_container_width=True)
'''
    #Criar Ficheiro Excel
    tabela_final.to_excel('tabela_final.xlsx', sheet_name='Test', index=False,float_format="%.5f",startrow=9, startcol=1)

    #Formatar Ficheiro Excel
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.styles import colors
    from openpyxl.styles.borders import Border, Side
    from openpyxl.cell import Cell
    from openpyxl.writer.excel import save_virtual_workbook

    from openpyxl.chart import (
        LineChart,
        BarChart,
        ScatterChart,
        Reference,
        Series,
    )

    wb = openpyxl.load_workbook('tabela_final.xlsx')

    sheet = wb.active

    sheet.merge_cells('B4:G5')
    sheet.merge_cells('B9:C9')
    sheet.merge_cells('D9:E9')
    sheet.merge_cells('F9:G9')

    fontObj1 = Font(bold=True)
    fontObj2 = Font(size=16, bold=True)

    sheet['B4'] = 'Tensile Testing - Raw Data for ' + str(sample_name)

    sheet['J10'] = 'Max. Displacement (mm)'
    sheet['J10'].font = fontObj1
    sheet['M10'] = tabela_final['Displacement'].max()

    sheet['J11'] = 'Max. Force (N)'
    sheet['J11'].font = fontObj1
    sheet['M11'] = tabela_final['Force'].max()

    sheet['J12'] = 'Tensile stress at Maximum Force (MPa)'
    sheet['J12'].font = fontObj1
    sheet['M12'] = tabela_final['Force'].max()/area

    sheet['B9'] = 'INSTRON Data'
    sheet['D9'] = 'DIC Data'
    sheet['F9'] = 'Calculation'

    sheet['B7'] = 'Young Modulus'
    sheet['C7'] = round(young_variable.slope/1000,4)
    sheet['D7'] = 'GPa'

    sheet['F7'] = 'Poissons Ratio'
    sheet['G7'] = abs(round(poisson_variable.slope,4))

    blueFill=PatternFill(fgColor="A5EAEF", fill_type = "solid")
    whiteFill=PatternFill(fgColor="ffffff", fill_type = "solid")

    for row in sheet['A1:Y{}'.format(sheet.max_row)]:
        for cell in row:
            cell.fill = whiteFill
            cell.alignment = Alignment(horizontal='center',vertical='center')


    sheet['B4'].fill = blueFill
    sheet['B4'].font = fontObj2

    if logo_f4y:
        img_f4y = openpyxl.drawing.image.Image('logo_f4y.png')
        img_f4y.anchor = 'B2'
        sheet.add_image(img_f4y)

    if logo_inegi:
        img_inegi = openpyxl.drawing.image.Image('logo_inegi.png')
        img_inegi.anchor = 'G2'
        sheet.add_image(img_inegi)

    sheet.row_dimensions[2].height = 20
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 16
    sheet.column_dimensions['E'].width = 16
    sheet.column_dimensions['F'].width = 16
    sheet.column_dimensions['G'].width = 16

    if plot_grafico:
        c1 = ScatterChart()
        c1.title = 'Tensile Stress vs Strain'

        xvalues = Reference(sheet, min_col=7, min_row=11, max_row=sheet.max_row)
        values = Reference(sheet, min_col=4, min_row=11, max_row=sheet.max_row)
        series = Series(values, xvalues, title_from_data=False)
        c1.series.append(series)

        s1 = c1.series[0]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "04939e" # Alterar cor
        s1.marker.graphicalProperties.line.solidFill = "04939e" # Alterar cor
        s1.graphicalProperties.line.noFill = True  # Esconder Linhas
        c1.style = 1
        c1.x_axis.title = 'Strain'
        c1.y_axis.title = 'Tensile Stress (MPa)'
        c1.y_axis.majorGridlines = None
        c1.x_axis.majorGridlines = None
        c1.x_axis.scaling.min = 0
        c1.y_axis.scaling.min = 0
        c1.height = 13
        c1.width = 30

        sheet.add_chart(c1, 'H14')
        
        st.balloons()
    streamtest = save_virtual_workbook(wb)
    st.download_button("Download Final Excel File",streamtest,'Tabela_Final_'+str(sample_name)+'.xlsx')
'''
